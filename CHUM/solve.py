from pysat.solvers import Solver
from pysat.card import *
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import pathlib
from datetime import date, datetime

parent_path = pathlib.Path(__file__).parent.resolve()

n_care_persons = 7 # 1 <= agent_index <= n_care_persons
n_shifts = 28 # 0 <= shift_index <= n_shifts-1
top_id = n_care_persons * n_shifts


def binary_variable_encoding(shift_index, agent_index):
    # 1 <= agent_index <= n_care_persons
    # 0 <= shift_index <= n_shifts-1
    binary_variable_index = shift_index * n_care_persons + agent_index
    # 1 <= binary_variable_index <= n_care_persons*n_shifts
    return binary_variable_index

def binary_variable_decoding(binary_variable_index):
    # 1 <= binary_variable_index <= n_care_persons*n_shifts
    shift_index = (binary_variable_index - 1) // n_care_persons
    agent_index = binary_variable_index - shift_index * n_care_persons
    return shift_index, agent_index

def service_constraint(shift_index):
    # use top_id to ensure that auxiliary variables are given new indices
    cnfplus = CNFPlus()
    if shift_index % 2 == 0: # day shift
        nb_staff = 3
    else: # night shift
        nb_staff = 1
    literals = [binary_variable_encoding(shift_index, agent_index) for agent_index in range(1, n_care_persons+1)]
    # print("service constraints literals:", literals)
    cnfplus.append([literals, nb_staff], is_atmost=True) # = CardEnc.equals(lits=literals, bound=nb_staff)
    cnfplus.append([[-x for x in literals], len(literals) - nb_staff], is_atmost=True) # implements an at_least constraitn
    # overall we have implemented an is_equal constraint 
    return cnfplus

def sliding_window_constraint(agent_index, sliding_window_size, bound):
    # print("new sw constraint")
    cnfplus = CNFPlus()
    for first_shift_index in range(n_shifts-sliding_window_size+1):
        # print('fsi:', first_shift_index)
        literals = [binary_variable_encoding(first_shift_index + t, agent_index) for t in range(sliding_window_size) if first_shift_index + t <= n_shifts-1]
        # print("sliding window literals:", literals)
        cnfplus.append([literals, bound], is_atmost=True) # cnfplus.atmost(literals, bound, top_id) # cnfplus.append(CardEnc.equals(lits=literals, bound=bound))
    return cnfplus

def days_between(d1, d2):
    d1 = datetime.strptime(d1, "%Y-%m-%d")
    d2 = datetime.strptime(d2, "%Y-%m-%d")
    return abs((d2 - d1).days)

def retrieve_individual_constraints(filename):

    person_index = {'Alice' : 1, 'Bob' : 2, 'Charlie' : 3, 'David' : 4, 'Eve' : 5, 'Fred' : 6, 'Gael' : 7}
    individual_constraints = []

    with open(parent_path / filename, 'r') as file:
        for line in file:
            person_talking, calendar_or_relative, absence_date, day_or_night = line.split(sep=', ')
            day_or_night = day_or_night.strip()
            agent_index = person_index[person_talking]

            if calendar_or_relative == 'calendrier':
                today = str(date.today())
                h24_index = days_between(today, absence_date)
                h12_index = h24_index * 2
            elif calendar_or_relative == 'relatif':
                relative_date = { 'demain' : 1, 'après-demain' : 2, 'la semaine prochaine' : 7, 'le mois prochain' : 30 }
                h24_index = relative_date[absence_date]
                h12_index = h24_index * 2
            else:
                print('Error: calendar_or_relative should be either calendrier ou relatif')

            if day_or_night == 'nuit':
                h12_index += 1
            elif day_or_night == 'jour':
                pass
            else:
                print('Error: day_or_night should be either jour or nuit')

            binary_variable_index = binary_variable_encoding(h12_index, agent_index)
            individual_constraints.append([-binary_variable_index])

    return individual_constraints

def compute_schedule():

    formula = CNFPlus()

    print('retrieving global constraints..')
    for shift_index in range(n_shifts):
        formula.extend(service_constraint(shift_index))

    for agent_index in range(1, n_care_persons+1):
        formula.extend(sliding_window_constraint(agent_index, 14, 4)) # no more than 4 shifts in a week
        formula.extend(sliding_window_constraint(agent_index, 7, 3)) # no more than 3 day-consecutive or night-consecutive shifts
        formula.extend(sliding_window_constraint(agent_index, 2, 1)) # no 24h in a row shifts

    print('retrieving individual constraints..')
    individual_constraints = retrieve_individual_constraints('individual_constraints.txt')
    formula.extend(individual_constraints)
    # print(formula.clauses)

    solver = Solver("minicard")
    solver.append_formula(formula)
    print('computing a solution schedule..')
    solver.solve()
    model = solver.get_model()
    # print('model:', model)

    return model

def write_to_excel(model):

    print('writing schedule to excel file')
    suffix = "schedule.xlsx"
    filename = parent_path / suffix

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    red_index = 2
    redFill = PatternFill(start_color=colors.COLOR_INDEX[red_index],
                   end_color=colors.COLOR_INDEX[red_index],
                   fill_type='solid')
    green_index = 3
    greenFill = PatternFill(start_color=colors.COLOR_INDEX[green_index] ,
                   end_color=colors.COLOR_INDEX[green_index],
                   fill_type='solid')
    

    for v in model:
        shift_index, agent_index = binary_variable_decoding(abs(v))
        value = v > 0
        cell = ws.cell(row=agent_index+1, column=shift_index+2)
        if value:
            cell.fill = greenFill 
        else:
            cell.fill = redFill

    
    individual_constraints = retrieve_individual_constraints('individual_constraints.txt')
    for binary_variable_index_list in individual_constraints:
        shift_index, agent_index = binary_variable_decoding(-binary_variable_index_list[0])
        ws.cell(row=agent_index+1, column=shift_index+2, value='no')


    wb.save(filename)

print()
model = compute_schedule()
print(model)
#write_to_excel(model)
print()







