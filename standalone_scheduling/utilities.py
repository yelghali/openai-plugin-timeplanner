#module imports
import datetime
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
import pathlib

# local imports
import cosmos

parent_path = pathlib.Path(__file__).parent.resolve()

def binary_variable_encoding(shift_index, staff_index, n_staff):
    # 1 <=_staff_index <= n_staff
    # 0 <= shift_index <= n_shifts-1
    binary_variable_index = shift_index * n_staff + staff_index
    # 1 <= binary_variable_index <= n_staff*n_shifts
    return binary_variable_index

def binary_variable_decoding(binary_variable_index, n_staff):
    # 1 <= binary_variable_index <= n_staff*n_shifts
    shift_index = (binary_variable_index - 1) // n_staff
    staff_index = binary_variable_index - shift_index * n_staff
    return shift_index, staff_index

def days_between(d1, d2):
    d1 = datetime.strptime(d1, "%Y-%m-%d")
    d2 = datetime.strptime(d2, "%Y-%m-%d")
    return abs((d2 - d1).days)

def write_to_excel(model, constraints, suffix, n_staff):

    # print('writing schedule to excel file')
    # suffix = "schedule.xlsx"
    filename = parent_path / suffix

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    red_index = 2
    redFill = PatternFill(start_color=colors.COLOR_INDEX[red_index],
                   end_color=colors.COLOR_INDEX[red_index],
                   fill_type='solid')
    green_index = 3
    greenFill = PatternFill(start_color=colors.COLOR_INDEX[green_index] ,
                   end_color=colors.COLOR_INDEX[green_index],
                   fill_type='solid')
    
    for v in model:
        shift_index, staff_index = binary_variable_decoding(abs(v), n_staff)
        value = v > 0
        cell = ws.cell(row=staff_index+1, column=shift_index+2)
        if value:
            cell.fill = greenFill 
        else:
            cell.fill = redFill
    
    for binary_variable_index_list in constraints:
        shift_index, staff_index = binary_variable_decoding(-binary_variable_index_list[0], n_staff)
        # print('shift_index:', shift_index)
        # print('staff_index:', staff_index)
        ws.cell(row=staff_index+1, column=shift_index+2, value='no')

    wb.save(filename)

def compute_distance(model_A, model_B):
    assert(len(model_A)==len(model_B))
    d = 0
    for v, w in zip(model_A, model_B):
        if v * w < 0: # w and w have opposite sign when they disagree
            d += 1
    return(d)

def date_and_time_as_string(shift_index):
    days = shift_index // 2
    today = date.today()
    calendar_date = str(today + timedelta(days=days))
    if shift_index%2==0:
        time = 'day'
    else:
        time = 'night'
    return calendar_date, time

def model_as_schedule(model, n_staff, staff_inverted_dict):
    schedule = []
    for v in model:
        shift_index, staff_index = binary_variable_decoding(abs(v), n_staff)
        if v > 0:
            calendar_date, time = date_and_time_as_string(shift_index)
            schedule.append({"id": str(abs(v)), "staff_name": staff_inverted_dict[str(staff_index)], "date": calendar_date, "time": time})
    return schedule

def schedule_as_model(schedule, n_staff, n_shifts, staff_dict):
    
    # print('schedule:', schedule)
    model = [-i for i in range(1, n_staff*n_shifts+1)]
    # print('model:', model)
    for item in schedule:
        staff_index = staff_dict[item["staff_name"]]
        shift_index = days_between(item["date"], str(date.today())) * 2
        if item["time"] == "night":
            shift_index += 1
        binary_variable_index = binary_variable_encoding(shift_index, staff_index, n_staff)
        model[binary_variable_index-1] = binary_variable_index
    print('model:', model)

def compute_binary_variable_index(staff_name, date, time):
    staff_index = staff_dict[staff_name]
    shift_index = days_between(date, str(date.today())) * 2
    if time == "night":
        shift_index += 1
    binary_variable_index = binary_variable_encoding(shift_index, staff_index, n_staff)
    return binary_variable_index

if __name__=='__main__':

    n_staff, n_shifts = 5, 4
    staff_dict = {'Alice' : 1, 'Bob' : 2, 'Charlie' : 3, 'David' : 4, 'Eve' : 5}

    schedule_as_model(cosmos.read('schedule') , n_staff, n_shifts)







